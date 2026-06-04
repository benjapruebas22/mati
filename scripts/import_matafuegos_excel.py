import argparse
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parents[1]
DB_PATH = BASE_DIR / "mpd.db"


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%Y", "%m/%y", "%m-%d-%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            if "%d" not in fmt:
                return date(parsed.year, parsed.month, 1).isoformat()
            return parsed.date().isoformat()
        except ValueError:
            pass

    match = re.fullmatch(r"(\d{1,2})/(\d{2,4})", text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if year < 100:
            year += 2000
        return date(year, month, 1).isoformat()

    return None


def parse_capacity(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", ".")
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None

    if number <= 0:
        return None
    if number > 1000:
        number = number / 100000.0
    return round(number, 2)


def normalize_sede(raw):
    text = str(raw or "").strip().upper()
    if not text:
        return "", ""

    text = text.replace("_", " ")
    match = re.search(r"S\s*-?\s*(\d{1,2})", text)
    sede = f"S{int(match.group(1)):02d}" if match else text.split()[0].replace("-", "")

    piso = ""
    if "PB" in text:
        piso = "PB"
    elif re.search(r"\bP\s*1\b|\b1P\b|\bP1\b", text):
        piso = "P1"
    elif re.search(r"\bP\s*2\b|\b2P\b|\bP2\b", text):
        piso = "P2"
    elif re.search(r"\bP\s*3\b|\b3P\b|\bP3\b", text):
        piso = "P3"

    return sede, piso


def infer_estado(fecha_vencimiento):
    if not fecha_vencimiento:
        return "Sin fecha"
    try:
        venc = datetime.fromisoformat(fecha_vencimiento).date()
    except ValueError:
        return "Fecha invalida"

    today = date.today()
    delta = (venc - today).days
    if delta < 0:
        return "Vencido"
    if delta == 0:
        return "Vence hoy"
    if delta <= 45:
        return "Vence <=45d"
    return "Vigente"


def main():
    parser = argparse.ArgumentParser(description="Importa matafuegos desde un Excel a mpd.db")
    parser.add_argument("xlsx", help="Ruta al archivo Excel")
    parser.add_argument("--sheet", default="Base de Datos ", help="Nombre de la hoja fuente")
    parser.add_argument("--db", default=str(DB_PATH), help="Ruta de la base SQLite")
    parser.add_argument("--keep-existing", action="store_true", help="No borra los matafuegos previos")
    args = parser.parse_args()

    db_path = Path(args.db)
    xlsx_path = Path(args.xlsx)

    if not db_path.exists():
        raise SystemExit(f"No se encontró la base de datos: {db_path}")
    if not xlsx_path.exists():
        raise SystemExit(f"No se encontró el Excel: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"No existe la hoja '{args.sheet}'. Hojas: {', '.join(wb.sheetnames)}")

    ws = wb[args.sheet]

    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS matafuegos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cod_sede TEXT NOT NULL DEFAULT '',
            sede TEXT NOT NULL,
            piso TEXT,
            local TEXT,
            tipo TEXT NOT NULL DEFAULT '',
            capacidad_kg REAL,
            numero_serie TEXT,
            ubicacion TEXT,
            fecha_recarga TEXT,
            fecha_vencimiento TEXT,
            fecha_prueba_hidro TEXT,
            estado TEXT DEFAULT 'Sin dato',
            activo INTEGER DEFAULT 1,
            lote_vencimiento TEXT,
            observaciones TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT
        )
    """)

    if not args.keep_existing:
        cur.execute("DELETE FROM matafuegos")
        cur.execute("DELETE FROM sqlite_sequence WHERE name='matafuegos'")

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_sede = row[0]
        numero_serie = str(row[1] or "").strip()
        ubicacion = str(row[2] or "").strip()
        capacidad_kg = parse_capacity(row[3])
        fecha_recarga = parse_date(row[4])
        fecha_vencimiento = parse_date(row[5])
        observaciones_raw = row[6]
        observaciones = "" if observaciones_raw is None else str(observaciones_raw).strip()

        if not str(cod_sede or "").strip() or not numero_serie:
            skipped += 1
            continue

        sede, piso = normalize_sede(cod_sede)
        lote_vencimiento = ""
        if fecha_vencimiento:
            venc = datetime.fromisoformat(fecha_vencimiento).date()
            lote_vencimiento = f"{venc.month:02d}/{venc.year}"

        estado = infer_estado(fecha_vencimiento)
        tipo = "Matafuego"

        cur.execute(
            """
            INSERT INTO matafuegos(
                cod_sede, sede, piso, local, tipo, capacidad_kg, numero_serie,
                ubicacion, fecha_recarga, fecha_vencimiento, fecha_prueba_hidro,
                estado, activo, lote_vencimiento, observaciones, created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now','localtime'),datetime('now','localtime'))
            """,
            (
                sede,
                sede,
                piso,
                "",
                tipo,
                capacidad_kg,
                numero_serie,
                ubicacion,
                fecha_recarga,
                fecha_vencimiento,
                None,
                estado,
                1,
                lote_vencimiento,
                observaciones,
            ),
        )
        inserted += 1

    con.commit()
    con.close()

    print(f"Insertados: {inserted}")
    print(f"Omitidos: {skipped}")


if __name__ == "__main__":
    main()
